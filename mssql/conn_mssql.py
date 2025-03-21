import tkinter as tk
from tkinter import filedialog
from tkinter import *
from openpyxl import load_workbook
import pyodbc
from PIL import ImageTk


def open_file():
    file_path = filedialog.askopenfilename()
    if file_path:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index > 0:
                filtered_row = [cell for idx, cell in enumerate(row) if idx != 2 and cell is not None]
                if filtered_row:
                    conn_db(filtered_row)

def conn_db(update_number : list ):
    # 連到 SQL Server
    DB_DRIVER = '{SQL Server}' # DB種類
    SERVER = ''
    DATABASE = '' #修改哪個 table
    UID = ''  
    PWD = ''

    conn = pyodbc.connect(f'DRIVER={DB_DRIVER};SERVER={SERVER};DATABASE={DATABASE};UID={UID};PWD={PWD}')

    # 自動執行script
    conn.autocommit = True
    # 創建執行對象
    cursor = conn.cursor()


    for count in range(len(update_number)) :

        # 執行 script
        try :
            cursor.execute("""
            DECLARE @Result int
            DECLARE @RtnValue VARCHAR(50)
            DECLARE @UserID INT
            DECLARE @UpperUserID INT
            SET @UserID = {UsrID}
            SET @UpperUserID = {UpperUserID}
            EXEC Backend.ChangeUpperAgent
                @UserID  = @UserID,
                @UpperUserID = @UpperUserID,
                @Result  = @Result OUTPUT,
                @RtnValue = @RtnValue OUTPUT
            SELECT @Result, @RtnValue
            """.format(UsrID=update_number[count] ,UpperUserID=update_number[count+1] ))

            # 獲取輸出參數
            result = cursor.fetchall()


            # 關閉DB
            cursor.close()
            conn.close()

        except:
            result_text.insert('end', f'下級{update_number[count]} 或新上級{update_number[count+1]} 的ID錯誤, 請重新確認' + '\n')

        else:
            for row in result:
                try :
                    Result = row[0]
                    RtnValue = row[1]
                    if Result == 1 :
                        result_text.insert('end', f'{update_number[count]} {RtnValue} + \n')
                    else:
                        raise Exception
                except:
                    result_text.insert('end', f'{update_number[count]} {RtnValue} + \n')
        break



# 主窗口
window = tk.Tk()
window.title('綁定名單修改')
window.geometry("500x650")
window.resizable(False,False)

maple_img = ImageTk.PhotoImage(file="./img/MapleStory.ico")
window.iconphoto(True , maple_img)

img = ImageTk.PhotoImage(file="./img/maplestory_tittle.png")
canvas = Canvas(width=280 , height=209)
canvas.create_image(150 , 100 , image=img )
canvas.pack(pady=10)

message = Label(text="正式環境 (需撥vpn)" , font=("標楷體" , 20))
message.pack(side="top" )

# 添加按鈕，點擊按鈕選擇文件並讀取内容
button = tk.Button(window, text="選擇檔案(.xlsx結尾)", font=("標楷體" , 15) , bg="#02DF82" ,command=open_file )
button.pack(pady=10)

result_text = Text(window)
result_text.pack(pady=20)

# 運行循環
window.mainloop()

